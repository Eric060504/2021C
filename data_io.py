"""Excel 输入清洗、结果表构建及模板写入工具。"""
from __future__ import annotations

from pathlib import Path
import shutil
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from config import (
    CARRIER_XLSX,
    FINAL_ORDER_RESULT_XLSX,
    FINAL_TRANSPORT_RESULT_XLSX,
    ORDER_TEMPLATE_XLSX,
    SUPPLIER_XLSX,
    TRANSPORT_TEMPLATE_XLSX,
    WEEKS_HISTORY,
    WEEKS_PLAN,
)

SUPPLIER_ID = "供应商ID"
MATERIAL = "材料分类"
ORDER_SHEET = "企业的订货量（m³）"
SUPPLY_SHEET = "供应商的供货量（m³）"
LOSS_SHEET = "运输损耗率（%）"


# 生成历史工作表的标准周列名，确保不同输入表按同一周次对齐。
def week_columns(prefix: str = "W", weeks: int = WEEKS_HISTORY) -> list[str]:
    return [f"{prefix}{week:03d}" for week in range(1, weeks + 1)]


# 生成附件结果模板所要求的中文周标签。
def output_week_labels(weeks: int = WEEKS_PLAN) -> list[str]:
    return [f"第{week:02d}周" for week in range(1, weeks + 1)]


# 在逐行建模前确认订货表、供货表的供应商和材料分类完全对应。
def validate_supplier_frames(order: pd.DataFrame, supply: pd.DataFrame) -> None:
    required = {SUPPLIER_ID, MATERIAL}
    if not required <= set(order.columns) or not required <= set(supply.columns):
        raise ValueError("输入表必须具有供应商ID和材料分类列")
    if order[SUPPLIER_ID].tolist() != supply[SUPPLIER_ID].tolist():
        raise ValueError("订货量和供货量的供应商ID不一致")
    if order[MATERIAL].tolist() != supply[MATERIAL].tolist():
        raise ValueError("订货量和供货量的材料分类不一致")


# 读取供应商历史表，并统一清洗标识字段和周度数值字段。
def _read_supplier_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    frame = pd.read_excel(path, sheet_name=sheet_name)
    frame = frame.dropna(subset=[SUPPLIER_ID]).copy()
    frame[SUPPLIER_ID] = frame[SUPPLIER_ID].astype(str).str.strip()
    frame[MATERIAL] = frame[MATERIAL].astype(str).str.strip()
    cols = [SUPPLIER_ID, MATERIAL, *week_columns()]
    missing = set(cols) - set(frame.columns)
    if missing:
        raise ValueError(f"{sheet_name} 缺少列: {sorted(missing)}")
    # 仅保留模型需要的字段，避免额外说明列干扰后续计算。
    frame = frame.loc[:, cols]
    frame.loc[:, week_columns()] = frame.loc[:, week_columns()].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    return frame


# 统一读取两份题目附件，并在进入模型前完成结构校验。
def read_input_data(
    supplier_path: Path = SUPPLIER_XLSX,
    carrier_path: Path = CARRIER_XLSX,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """读取两份历史输入工作簿，并完成字段和行顺序校验。"""
    order = _read_supplier_sheet(Path(supplier_path), ORDER_SHEET)
    supply = _read_supplier_sheet(Path(supplier_path), SUPPLY_SHEET)
    validate_supplier_frames(order, supply)

    loss = pd.read_excel(Path(carrier_path), sheet_name=LOSS_SHEET).dropna(subset=["转运商ID"]).copy()
    loss["转运商ID"] = loss["转运商ID"].astype(str).str.strip()
    loss_cols = week_columns()
    missing = set(loss_cols) - set(loss.columns)
    if missing:
        raise ValueError(f"运输损耗率表缺少列: {sorted(missing)}")
    loss = loss.loc[:, ["转运商ID", *loss_cols]]
    loss.loc[:, loss_cols] = loss.loc[:, loss_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    return order, supply, loss


# 按附件 A 的完整供应商顺序展开订购结果；无订购量保持为空白。
def build_order_result_frame(plan: pd.DataFrame, supplier_ids: Iterable[str]) -> pd.DataFrame:
    ids = list(supplier_ids)
    result = pd.DataFrame(index=ids)
    result.index.name = SUPPLIER_ID
    for week, label in enumerate(output_week_labels(), start=1):
        values = plan[week] if week in plan.columns else pd.Series(0.0, index=plan.index, dtype=float)
        aligned = pd.to_numeric(values.reindex(ids), errors="coerce").fillna(0.0)
        # 模板以空白表示本周不订货，而非用数值零填充。
        result[label] = aligned.mask(aligned <= 1e-8)
    return result


# 按“规划周次 × 转运商”多级列结构展开附件 B 所需的转运结果。
def build_transport_result_frame(
    plan: dict[int, pd.DataFrame], supplier_ids: Iterable[str], carrier_ids: Iterable[str]
) -> pd.DataFrame:
    ids = list(supplier_ids)
    carriers = list(carrier_ids)
    values_by_column: dict[tuple[int, str], pd.Series] = {}
    for week in range(1, WEEKS_PLAN + 1):
        allocation = plan.get(week, pd.DataFrame(0.0, index=ids, columns=carriers))
        aligned_allocation = allocation.reindex(index=ids, columns=carriers, fill_value=0.0)
        for carrier in carriers:
            values = pd.to_numeric(aligned_allocation[carrier], errors="coerce").fillna(0.0)
            values_by_column[(week, carrier)] = values.mask(values <= 1e-8)
    result = pd.DataFrame(values_by_column, index=ids)
    result.index.name = SUPPLIER_ID
    result.columns = pd.MultiIndex.from_tuples(result.columns, names=["week", "carrier"])
    return result


# 首次输出时复制题目模板，避免直接修改原始附件。
def copy_template(template_path: Path, output_path: Path) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)


# 保存模板说明区和表头快照，用于写入后的结构一致性检查。
def _sheet_header_snapshot(ws, max_header_row: int = 6) -> list[list[object]]:
    return [[ws.cell(row, column).value for column in range(1, ws.max_column + 1)] for row in range(1, max_header_row + 1)]


# 保存供应商 ID 区域快照，防止填表时错位或覆盖供应商行。
def _supplier_id_snapshot(ws) -> list[object]:
    return [ws.cell(row, 1).value for row in range(7, ws.max_row + 1)]


# 对比输入模板与结果文件，确保程序仅填数值而未更改模板结构。
def _assert_same_template_structure(source_path: Path, output_path: Path, sheet_name: str) -> None:
    source_wb = load_workbook(source_path, read_only=False, data_only=False)
    output_wb = load_workbook(output_path, read_only=False, data_only=False)
    source_ws, output_ws = source_wb[sheet_name], output_wb[sheet_name]
    if _sheet_header_snapshot(source_ws) != _sheet_header_snapshot(output_ws):
        raise AssertionError(f"{sheet_name} 的表头或说明区域被修改")
    if _supplier_id_snapshot(source_ws) != _supplier_id_snapshot(output_ws):
        raise AssertionError(f"{sheet_name} 的供应商ID区域被修改")


# 将指定问题的订货结果写入附件 A 对应工作表。
def write_order_template(output_path: Path, sheet_name: str, result: pd.DataFrame) -> None:
    output_path = Path(output_path)
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    labels = output_week_labels()
    for row in range(7, ws.max_row + 1):
        supplier_id = ws.cell(row, 1).value
        if supplier_id not in result.index:
            continue
        for offset, label in enumerate(labels, start=2):
            value = result.loc[supplier_id, label]
            ws.cell(row, offset).value = None if pd.isna(value) else float(value)
    wb.save(output_path)


# 将指定问题的转运分配写入附件 B 对应工作表。
def write_transport_template(output_path: Path, sheet_name: str, result: pd.DataFrame) -> None:
    output_path = Path(output_path)
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    for row in range(7, ws.max_row + 1):
        supplier_id = ws.cell(row, 1).value
        if supplier_id not in result.index:
            continue
        for week in range(1, WEEKS_PLAN + 1):
            # 每周固定对应 8 家转运商，名称从模板表头读取以避免错列。
            for carrier_index in range(8):
                column = 2 + (week - 1) * 8 + carrier_index
                carrier = ws.cell(6, column).value
                value = result.loc[supplier_id, (week, carrier)]
                ws.cell(row, column).value = None if pd.isna(value) else float(value)
    wb.save(output_path)


# 将问题 2、3、4 的结果增量写入同一对最终附件，保留其他工作表结果。
def write_combined_question_workbooks(
    question: int,
    order_result: pd.DataFrame,
    transport_result: pd.DataFrame,
    order_path: Path = FINAL_ORDER_RESULT_XLSX,
    transport_path: Path = FINAL_TRANSPORT_RESULT_XLSX,
) -> tuple[Path, Path]:
    """将一个问题的结果写入共享的最终附件。

    两份模板均预置问题 2--4 的结果工作表：首次调用时复制模板，
    后续仅更新当前问题的工作表，从而保留其他问题已写入的结果。
    """
    if question not in (2, 3, 4):
        raise ValueError("联合结果附件仅包含问题2、问题3和问题4")

    order_path = Path(order_path)
    transport_path = Path(transport_path)
    if not order_path.exists():
        copy_template(ORDER_TEMPLATE_XLSX, order_path)
    if not transport_path.exists():
        copy_template(TRANSPORT_TEMPLATE_XLSX, transport_path)

    order_sheet = f"问题{question}的订购方案结果"
    transport_sheet = f"问题{question}的转运方案结果"
    write_order_template(order_path, order_sheet, order_result)
    write_transport_template(transport_path, transport_sheet, transport_result)
    _assert_same_template_structure(ORDER_TEMPLATE_XLSX, order_path, order_sheet)
    _assert_same_template_structure(TRANSPORT_TEMPLATE_XLSX, transport_path, transport_sheet)
    return order_path, transport_path


# 提供给测试调用的模板完整性断言入口。
def assert_template_intact(question: int, order_path: Path, transport_path: Path) -> None:
    _assert_same_template_structure(ORDER_TEMPLATE_XLSX, order_path, f"问题{question}的订购方案结果")
    _assert_same_template_structure(TRANSPORT_TEMPLATE_XLSX, transport_path, f"问题{question}的转运方案结果")
